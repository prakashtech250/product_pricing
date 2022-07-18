import requests
from scrapy.selector import Selector
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys
from tkinter import *
from tkinter import ttk
from threading import *
from tkinter import messagebox
import urllib3
from time import sleep
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

WAIT = 50
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

HEADERS = {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36'
        }

class productPricing:   
    def open_browser(self):
        tbox_msg = "Opening google chrome to get data..."                     
        TBox.insert(END, tbox_msg)
        TBox.yview(END)
        options = Options()
        options.add_argument("--incognito")
        options.add_argument("--headless")
        options.add_experimental_option('excludeSwitches', ['enable-logging'])

        exec_path = '{}/driver/chromedriver.exe'.format(os.getcwd())
        try:
            options.binary_location = r'C:/Program Files (x86)/Google/Chrome/Application/chrome.exe'
            self.driver = webdriver.Chrome(options=options,executable_path=exec_path)
            print('using 32 bit chrome')
        except:
            options.binary_location = r'C:/Program Files/Google/Chrome/Application/chrome.exe'
            self.driver = webdriver.Chrome(options=options,executable_path=exec_path)
            print('Using 64 bit chrome')

    def get_page_source(self):
        page_source = self.driver.page_source
        res = Selector(text=page_source)
        return res

    def cookiesToDict(self):
        data = None
        cookies = {}
        try:
            with open("cookies_bunnings.json","r") as f:
                data = json.load(f)
        except Exception as e:
            return None
        
        try:
            for i in data:
                cookies[i['name']] = i['value']
        except Exception as e:
            print(str(e))
            return None

        return cookies 

    def website_changed(self):
        self.TBox = Text(self.root)
        self.TBox.pack()
        website = self.websitechoosen.get()
        msg = f'You have selected {website}. Please wait!!! '
        self.TBox.insert(END,msg)

    def get_response(self,url , headers, cookies):
        while True:
            try:
                try:
                    response = requests.get(url, headers=headers, cookies=cookies)
                except:
                    response = requests.get(url, headers=headers, cookies=cookies, verify=False) 
                print(response.status_code)
                return response.text
                break
            except:
                tbox_msg = 'Connection error. Please check your internet connection...'
                TBox.insert(END, tbox_msg)
                TBox.yview(END)
                sleep(5)
                pass
    
    def get_json(self, response):
        return json.loads(response) 

    def get_Selector(self, response):
        sel = Selector(text=response)
        return sel
    
    def create_filename(self, website):
        if not os.path.exists('output'):
            os.makedirs('output')
        now = datetime.now()
        dt_string = now.strftime("%Y%m%d_%H_%M")
        filename = 'output/{}_{}.xlsx'.format(website,dt_string)
        return filename
    
    def clean_data(self, data):
        output = []
        for d in data:
            if d:
                output.append(d.strip().replace('\n',' '))
            else:
                output.append('')
        return output

    def nzsafetyblackwoods(self):
        global running
        filename = self.create_filename('nzsafetyblackwoods')
        sheet_title = ['Manifacture Code','Sku','Title','Product Url','Price incl gst','Price excl gst','Brand']
        wb = Workbook()
        ws = wb.active
        ws.append(sheet_title)
        base_url = 'https://nzsafetyblackwoods.co.nz'
        
        count = 1
        
        url = 'https://nzsafetyblackwoods.co.nz/en/'
        catUrl = 'https://nzsafetyblackwoods.co.nz/catalog/popularcategoryitems?categoryId={}'
        productsUrl = 'https://nzsafetyblackwoods.co.nz/product/productbycategory?categoryId={}&count=50&pageNumber={}'
        response = self.get_Selector(self.get_response(url, HEADERS, None))
        primary_cats_index = [x.split('/')[-1].split('-')[-1] for x in response.css('.app-megamenu-item::attr(href)').extract()]
        primary_cats = [x.strip() for x in response.css('.app-megamenu-item::text').extract()]
        i = 0
        for primary_cat_index in primary_cats_index[i:]:
            if running:
                url = catUrl.format(primary_cat_index)
                response = self.get_json(self.get_response(url, HEADERS, None))
                sub_cats_index = [x.get('Id') for x in response]
                sub_cats = [x.get('Name') for x in response]
                primary_cat = primary_cats[i]
                i+=1

                j=0
                for sub_cat_index in sub_cats_index:
                    if running:
                        url = catUrl.format(sub_cat_index)
                        response = self.get_json(self.get_response(url, HEADERS, None))
                        subSub_cats_index = [x.get('Id') for x in response]
                        subSub_cats = [x.get('Name') for x in response]
                        sub_cat = sub_cats[j]
                        j+=1 

                        k = 0
                        for subSub_cat_index in subSub_cats_index[k:]:
                            page_no = 0
                            if running:
                                url = productsUrl.format(subSub_cat_index, page_no)
                                response = self.get_json(self.get_response(url, HEADERS, None))
                                total_pages = response.get('NumberOfPages')
                                total_items = response.get('TotalRecords')
                                subSub_cat = subSub_cats[k]
                                k+=1
                                tbox_msg = "{} > {} > {}, total items: {}".format(primary_cat, sub_cat, subSub_cat, total_items)
                                TBox.insert(END, '')
                                TBox.insert(END, tbox_msg)
                                TBox.yview(END)
                                while page_no <= total_pages and running:

                                    for json in response.get('ProductOverviewModels'):
                                        if running:
                                            productCode = ''
                                            brand = ''
                                            product_link = base_url + json.get('SeName')
                                            sku = json.get('Sku')
                                            title = json.get('Name')
                                            specialPrice = json.get('ProductPrice').get('PriceWithGST')
                                            retailPrice = json.get('ProductPrice').get('Price')
                                            description = json.get('FullDescription')
                                            sel = self.get_Selector(description)
                                            specTerms = sel.css('SpecTerm::text').extract()
                                            specDescs = sel.css('SpecDesc::text').extract()
                                            brand_index = 0
                                            for specTerm in specTerms:
                                                if 'Brand' in specTerm:
                                                    brand = specDescs[brand_index]
                                                    break
                                                else:
                                                    brand_index += 1
                                                    
                                            mfr_index = 0
                                            for specTerm in specTerms:
                                                if 'Mfr Part No' in specTerm or 'Mfr No' in specTerm:
                                                    productCode = specDescs[mfr_index]
                                                    break
                                                else:
                                                    mfr_index += 1
                                            try:
                                                data = [productCode,sku, title, product_link, specialPrice, retailPrice, brand]
                                                ws.append(data)
                                            except:
                                                title = json.get('DefaultPictureModel').get('Title')
                                                data = self.clean_data([productCode, sku, title, product_link, specialPrice, retailPrice, brand])
                                                try:
                                                    ws.append(data)
                                                except:
                                                    ws.append([sku, '','',productCode])
                    
                                            tbox_msg = "{}: {}".format(count, title)
                                            TBox.insert(END, tbox_msg)
                                            TBox.yview(END)
                                            count += 1
                                    wb.save(filename)
                                    page_no += 1
                                    url = productsUrl.format(subSub_cat_index, page_no)
                                    response = self.get_json(self.get_response(url, HEADERS, None))
                                
        wb.save(filename)
        TBox.insert(END,'Output is saved as {}'.format(filename))
        TBox.yview(END)
    
    def placemakers(self):
        global running
        filename = self.create_filename('placemakers')
        sheet_title = ['Product Code','Title','Product Url','Price','Brand']
        wb = Workbook()
        ws = wb.active
        ws.append(sheet_title)
        count = 1
        base_url = 'https://shop.placemakers.co.nz'
        cookies = {
            'loadedAlready': 'YES',
            'JSESSIONID': '68A8B187C92874A9438404FB7F01D836.accstorefront-857476864f-jg4zn',
            'TS0111fe1e': '0117e34ade5b1b30dfa2350728e3968e413598fdb4ea6dabedabb7c454cc9206be720db0649affcb02e783bb11db54b220296c7157',
            'anonymous-consents': '^%^5B^%^5D',
            'cookie-notification': 'NOT_ACCEPTED',
            'ROUTE': '.accstorefront-857476864f-jg4zn',
            '_gcl_au': '1.1.849293780.1622789327',
            '_rollupGa': 'GA1.3.1164923216.1622789328',
            '_rollupGa_gid': 'GA1.3.193969830.1622789328',
            '_ga': 'GA1.1.1164923216.1622789328',
            '_hjTLDTest': '1',
            '_hjid': 'c7ede076-bcad-420c-a247-ea1d5f9ace62',
            '_ga_NSNV9JBLYQ': 'GS1.1.1622789327.1.1.1622790700.59',
            'TS01ac0357': '0117e34ade62a688d1f9748386621d833b6f7e342b990972ee0e2cc7a628140c4e385ce4311eba3dde4e71c4eee54846c753042dee',
        }
        headers = {
            'Connection': 'keep-alive',
            'Cache-Control': 'max-age=0',
            'sec-ch-ua': '^\\^',
            'sec-ch-ua-mobile': '?0',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-User': '?1',
            'Sec-Fetch-Dest': 'document',
            'Referer': 'https://shop.placemakers.co.nz',
            'Accept-Language': 'en-US,en;q=0.9',
        }
        url = 'https://shop.placemakers.co.nz/fbb2cstorefront/placemakersB2C/en/NZD#'
        response = self.get_Selector(self.get_response(url, headers, None))
        catLinks = response.css('.dropdown-submenu .cat-level-one a::attr(href)').extract()
        categories = response.css('.dropdown-submenu .cat-level-one a::text').extract()
        i=0
        for catLink in catLinks:
            category = categories[i]
            i+=1 
            if running:
                TBox.insert(END, '')
                tbox_msg = "Category: {}".format(category)
                TBox.insert(END, tbox_msg)
                TBox.yview(END)
                url = base_url + catLink
                response = self.get_Selector(self.get_response(url, HEADERS, None))     
                product_div = response.css('.product-item')
                next_page = response.css('.pagination-next a::attr(href)').extract()
                for product in product_div:
                    productUrl =base_url + product.css('a::attr(href)').get()
                    title = product.css('.details .name::text').get().replace('\xa0','')
                    sku = product.css('.sku::text').get().replace('SKU:\xa0','')
                    price = product.css('.details .price::text').get()
                    available = product.css('.details .stock::text').get().replace('\xa0',' ')
                    brand = product.css('.details .manufacturer::text').get()
                    data = self.clean_data([sku, title,productUrl,price, brand])
                    print(data)
                    ws.append(data) 
                    tbox_msg = "{}: {}".format(count, data[1])
                    TBox.insert(END, tbox_msg)
                    TBox.yview(END)
                    count += 1                 
                wb.save(filename)
                while len(next_page) > 0 and running:
                    url = base_url + next_page[0]
                    response = self.get_Selector(self.get_response(url, headers, None))  
                    next_page = response.css('.pagination-next a::attr(href)').extract()  
                    product_div = response.css('.product-item')
                    for product in product_div:
                        productUrl =base_url + product.css('a::attr(href)').get()
                        title = product.css('.details .name::text').get().replace('\xa0',' ')
                        sku = product.css('.sku::text').get().replace('SKU:\xa0','')
                        price = product.css('.details .price::text').get()
                        available = product.css('.details .stock::text').get().replace('\xa0',' ')
                        brand = product.css('.details .manufacturer::text').get()
                        data = self.clean_data([sku, title,productUrl,price, brand])
                        print(data)
                        ws.append(data) 
                        tbox_msg = "{}: {}".format(count, data[1])
                        TBox.insert(END, tbox_msg)
                        TBox.yview(END)
                        count += 1 
                    wb.save(filename)    
        wb.save(filename)
        TBox.insert(END,'Output is saved as {}'.format(filename))
        TBox.yview(END)
    
    def hectorjones(self):
        global running
        filename = self.create_filename('hectorjones')
        sheet_title = ['Product Code','Title','Product Url','Special Price', 'Retail Price']
        wb = Workbook()
        ws = wb.active
        ws.append(sheet_title)
        count = 1
        base_url = 'https://www.hectorjones.co.nz'
        urllib3.disable_warnings()
        cookies = {
            '88687bc6becdfff6072b8a6afb67ca59': '250be79d42bc05ea140fcf740a5c9e22',
        }
        headers = {
            'Connection': 'keep-alive',
            'Cache-Control': 'max-age=0',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Sec-GPC': '1',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-User': '?1',
            'Sec-Fetch-Dest': 'document',
            'Accept-Language': 'en-US,en;q=0.9',
            'Referer': 'https://www.hectorjones.co.nz/shop.html'
        }
        url = 'https://www.hectorjones.co.nz/shop.html?start=0'
        response = self.get_Selector(self.get_response(url, headers, cookies))
        product_div = response.css('.product-container')
        next_page = response.css('.pagination-next a::attr(href)').extract()
        for div in product_div:
            if running:
                sku = div.css('.product-sku::text').get().split(':')[1]
                title = div.css('h2 a::text').get()
                product_url = base_url + div.css('.vm-details-button a::attr(href)').get()
                specialPrice = div.css('.PricesalesPrice::text').get()
                retailPrice = div.css('.vm-price-value .PricebasePriceWithTax::text').get()
                data = self.clean_data([sku, title, product_url ,specialPrice, retailPrice]) 
                ws.append(data)
                tbox_msg = "{}: {}".format(count, title)
                TBox.insert(END, tbox_msg)
                TBox.yview(END)
                count += 1
        wb.save(filename)            
        while len(next_page) > 0 and running:  
            url = base_url + next_page[0]
            response = self.get_Selector(self.get_response(url, headers, cookies))
            next_page = response.css('.pagination-next a::attr(href)').extract()
            product_div = response.css('.product-container')
            for div in product_div:
                if running:
                    sku = div.css('.product-sku::text').get().split(':')[1]
                    title = div.css('h2 a::text').get()
                    product_url = base_url + div.css('.vm-details-button a::attr(href)').get()
                    specialPrice = div.css('.PricesalesPrice::text').get()
                    retailPrice = div.css('.vm-price-value .PricebasePriceWithTax::text').get()
                    data = self.clean_data([sku, title, product_url ,specialPrice, retailPrice]) 
                    ws.append(data)
                    tbox_msg = "{}: {}".format(count, title)
                    TBox.insert(END, tbox_msg)
                    TBox.yview(END)
                    count += 1
            wb.save(filename)
        wb.save(filename)
        TBox.insert(END,'Output is saved as {}'.format(filename))
        TBox.yview(END)

    def bunnings2(self):
        headers = {
            'authority': 'www.bunnings.co.nz',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-language': 'en-US',
            'cache-control': 'max-age=0',
            # Requests sorts cookies= alphabetically
            # 'cookie': f"nodeServed=true; uSessionId=de58d170-064b-11ed-96f7-9f45dc0c3db8; personalization_session_id=de58d171-064b-11ed-96f7-9f45dc0c3db8; budp_nz#lang=en; ASP.NET_SessionId=e1134u0ukt1pinhnnzkkqyts; dtCookie=v_4_srv_4_sn_6A7892869C90D84AD4C55FCCA9E97D71_perc_100000_ol_0_mul_1_app-3A0adcc72d087158b6_0_rcs-3Acss_0; Bunnings.Platform=V2; Bunnings.Platform.Fallback=V2; __cf_bm=RTq1oA3FVz91eQ.QilGOwzvKppjnVOx8RtHGhMwIFok-1658115845-0-AQQrywKEMOizwyHh+ryy3t9RLgm9+azlpXLgj06WozmHD6W/4AtCXHnLiqJTLFi8VAd+gZxJa6euBqf7g5jQcrg=; rxVisitor=1658115844989ELR6A29SILUHIAP3EEN3R71AKE8ACNT7; defaultStoreID=9489; defaultRegionCode=NI_Zone_9; returningVisitor=true; recentSearches=Makita; ctz=-5.75; guest-token-storage={\"token\":\"eyJhbGciOiJSUzI1NiIsImtpZCI6IkJGRTFEMDBBRUZERkVDNzM4N0E1RUFFMzkxNjRFM0MwMUJBNzVDODciLCJ0eXAiOiJKV1QiLCJ4NXQiOiJ2LUhRQ3VfZjdIT0hwZXJqa1dUandCdW5YSWMifQ.eyJpc3MiOiJodHRwczovL2J1bm5pbmdzLmNvbS5hdS8iLCJuYmYiOjE2NTgxMTU4NTUsImlhdCI6MTY1ODExNTg1NSwiZXhwIjoxNjU4NTQ3ODU1LCJhdWQiOlsiQ2hlY2tvdXQtQXBpIiwiY3VzdG9tZXJfYnVubmluZ3MiLCJodHRwczovL2J1bm5pbmdzLmNvbS5hdS9yZXNvdXJjZXMiXSwic2NvcGUiOlsiY2hrOmV4ZWMiLCJjbTphY2Nlc3MiLCJlY29tOmFjY2VzcyIsImNoazpwdWIiXSwiYW1yIjpbImV4dGVybmFsIl0sImNsaWVudF9pZCI6ImJ1ZHBfZ3Vlc3RfdXNlcl9ueiIsInN1YiI6ImJlNWQ3NjhjLWYwNmEtNGNmOC1iNWMwLWIwMmE2ZjMyNTk3MiIsImF1dGhfdGltZSI6MTY1ODExNTg1NCwiaWRwIjoibG9jYWxsb29wYmFjayIsImItaWQiOiJiZTVkNzY4Yy1mMDZhLTRjZjgtYjVjMC1iMDJhNmYzMjU5NzIiLCJiLXJvbGUiOiJndWVzdCIsImItdHlwZSI6Imd1ZXN0IiwibG9jYWxlIjoiZW5fTloiLCJiLWNvdW50cnkiOiJOWiIsInVzZXJfbmFtZSI6ImJlNWQ3NjhjLWYwNmEtNGNmOC1iNWMwLWIwMmE2ZjMyNTk3MiIsImFjdGl2YXRpb25fc3RhdHVzIjoiRmFsc2UiLCJiLXJiYWMiOlt7ImFzYyI6IjQxYjJlNDY5LTI2ODUtNGMxMC1hMzlmLTUwZGE5MTIyZTVmNiIsInR5cGUiOiJDIiwicm9sIjpbIkNISzpHdWVzdCJdfV0sInNpZCI6IkVFOTlERjJGQTMxQzY4NzBGNENBMDBFODc5OThGNjJBIn0.RHHPpCdFTZJ--7UTAtyMNYru-7nYTZlk_KXIavA9jDmVZDFH_YUyVPHb7tPmlpRbbNmKP1DHaOdKt1uzJM8jan3ZgsvyA3XEQClqXlmDzoetMw0uKVNvPrdI-ZGnvZSVWRQQ4DjgxZDdaEJhIaIR-UjEQh8OUCyKD8uwAw0RScW10mkDBZQ1MtxWqI-PVXQ6X65dCSFazbEcp7drcLsI9LJEXwxrqCXnjN2yVV5fgX6XjU7m6phdUiNWlfaefZSvYMreMfrS9fgMMmk4I1ML926E82XRaIae9bSZDTCiukEFZ1Y53wZqs2DUtdLmbx2r0YXzIZnO8FAXS63RLZcTbQ\",\"expires\":1658547854900,\"s\":432000,\"clientToken\":true}; SC_ANALYTICS_GLOBAL_COOKIE=8a662160100a4d1da75937022f0c5c0f|False; origin_path=/search/products?q=Makita&sort=BoostOrder&page=1; rxvt=1658117658915|1658115844992; dtPC=4{315844985_229h-vRBEUUMMCAAEMLODEKKPKTKIAHKTLELAI-0e0;} dtLatC=1; AccessTokenCookie=%7B%22tokenvalue%22%3A%22be5d768c-f06a-4cf8-b5c0-b02a6f325972-1658115854%22%2C%22lastchecktime%22%3A%227%2F18%2F2022%203%3A44%3A16%20AM%22%7D; dtSa=-",
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-user': '?1',
            'sec-gpc': '1',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.53 Safari/537.36',
        }
        cookies = self.cookiesToDict()
        global running
        filename = self.create_filename('bunnings')
        sheet_title = ['Product Code','Title','Product Url','Price', 'Brand']
        wb = Workbook()
        ws = wb.active
        ws.append(sheet_title)
        params_url = 'https://www.bunnings.co.nz/search/products?q={}&brandname={}&sort=BoostOrder&pageSize=36&page={}'
        
        brands = ['Makita','Dewalt']
        base_url = 'https://www.bunnings.co.nz'
        count = 1

        def get_products_links(url):
            self.driver.get(url)   
            sleep(0.5)
            self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
            sleep(0.5)
            # res = self.get_response(url, headers, cookies)
            # response = self.get_Selector(res)
            response = self.get_page_source()
            links = []
            link_divs = response.css('article')
            for link_div in link_divs:
                link = link_div.css('a::attr(href)').extract()[0]  
                links.append(link)
            return links

        def get_product_details(response, url):    
            i = 0
            productCode = ''
            title = response.css('.productItemName::text').get()
            specHeading = response.css('.productSpecificationHeading::text').extract()
            spec = response.css('.productSpecificationHeading+ div::text').extract()
            for d in specHeading:
                if d == 'Model Number':
                    productCode = spec[i]
                    break
                else:
                    i+=1
            price = response.css('.productItemPrice p::text').extract()
            print(price)
            brand = response.css('.brandName::text').get()
            data = self.clean_data([productCode, title,url,price, brand])
            print(data)
            return data

        self.open_browser()
        for brand in brands:
            tbox_msg = "Brand: {}".format(brand)
            TBox.insert(END, '')
            TBox.insert(END, tbox_msg)
            page_no = 1
            while running:

                url = params_url.format(brand,brand,page_no)
                links = [base_url+l for l in get_products_links(url)]
                if len(links) == 0:
                    break
                else:
                    page_no +=1 
                for link in links:
                    if running:
                        self.driver.get(link)
                        try:
                            myElem = WebDriverWait(self.driver, 300).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.productItemName')))
                        except TimeoutException:
                            print("1: Loading took too much time!")
                        response = self.get_page_source()
                        # res = self.get_response(link, headers, cookies)
                        # response = self.get_Selector(res)
                        data = get_product_details(response, link)
                        # print(data)
                        ws.append(data)
                        tbox_msg = "{}: {}".format(count, data[1])
                        TBox.insert(END, tbox_msg)
                        TBox.yview(END)
                        count += 1
                wb.save(filename)
        wb.save(filename)
        TBox.insert(END,'Output is saved as {}'.format(filename))
        TBox.yview(END)
        self.driver.quit()               
                
    
    def bunnings(self):
        global running
        filename = self.create_filename('bunnings')
        sheet_title = ['Product Code','Title','Product Url','Price', 'Brand']
        wb = Workbook()
        ws = wb.active
        ws.append(sheet_title)
        count = 1
        next_page_params = '?pageSize=36&page={page_no}'
        self.open_browser()
        page_no = 1

        def get_products_links(url):
            all_links = []
            self.driver.get(url)   
            sleep(0.5)
            self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
            sleep(0.5)
            response = self.get_page_source()
            links = response.css('.productTileTitle::attr(href)').extract()    
            return links
        
        def get_product_details(response, url):
            
            i = 0
            productCode = ''
            title = response.css('..productItemName::text').get()
            specHeading = response.css('.productSpecificationHeading::text').extract()
            spec = response.css('.productSpecificationHeading+ div::text').extract()
            for d in specHeading:
                if d == 'Model Number':
                    productCode = spec[i]
                    break
                else:
                    i+=1
            price = response.css('.productPriceWrap p::text').extract()
            price = ''.join(price)
            brand = response.css('.brandName::text').get()
            data = self.clean_data([productCode, title,url,price, brand])
            return data


        base_url = 'https://www.bunnings.co.nz/products'
        url = 'https://api.prod.bunnings.com.au/v1/products/regionalCategory?depth=3'
        headers = {
            'Connection': 'keep-alive',
            'correlationid': '17f83490-bb8c-11eb-803a-e1b95e81bf30',
            'country': 'NZ',
            'currency': 'NZD',
            'isComposition': 'Y',
            'locale': 'en_NZ',
            'Authorization': 'Bearer eyJhbGciOiJSUzI1NiIsImtpZCI6IkJGRTFEMDBBRUZERkVDNzM4N0E1RUFFMzkxNjRFM0MwMUJBNzVDODciLCJ0eXAiOiJKV1QiLCJ4NXQiOiJ2LUhRQ3VfZjdIT0hwZXJqa1dUandCdW5YSWMifQ.eyJuYmYiOjE2MjE3MzEyNzEsImV4cCI6MTYyNzk1MjA3MSwiaXNzIjoiaHR0cHM6Ly9idW5uaW5ncy5jb20uYXUvIiwiYXVkIjpbImN1c3RvbWVyX2J1bm5pbmdzIiwiaHR0cHM6Ly9idW5uaW5ncy5jb20uYXUvcmVzb3VyY2VzIl0sImNsaWVudF9pZCI6ImJ1ZHBfc3NyX3JvYm90IiwicHJlZmVycmVkX3VzZXJuYW1lIjoiZ3Vlc3QiLCJiLXR5cGUiOiJndWVzdCIsImItaWQiOiJyb2JvdF9ndWVzdCIsInVzZXJfbmFtZSI6InJvYm90X2d1ZXN0IiwiaWF0IjoxNjIxNzMxMjcxLCJzY29wZSI6WyJlY29tOmFjY2VzcyJdfQ.KMOAc_NaujUy2GLrdz3ifQLwpFkvL7hyukkOXc75n8_5o29zgRiymQO8WL7esliEUDzywYTr_Cs5dyi782B68XKC9AKnnd3ev_czSUBRCmj8We0BkVheKsAD0PCpyly9cjyVnHVXgdtE9LCJile45s6uHpAs3Z11GjHQFm2sWETQQiysnKDGsRMv53JJURuuMlZbcdmjNe7Yq2xM0uEVQytNcxXSqH9Pikpo4yF6qTgi3RVJKlGq0dxGUviI1Olv4SmxvxzFkShceTvke4hUkbSCLHwq6m8Dt_jhCSSV94z9DbHtG3da0vfvD6GJdIudn0FrY-RLxIN5GSpz8O005Q',
            'userId': 'anonymous',
            'Accept': 'application/json, text/plain, */*',
            'clientId': 'mHPVWnzuBkrW7rmt56XGwKkb5Gp9BJMk',
            'locationCode': '9489',
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36',
            'X-region': 'NI_Zone_9',
            'sessionid': '9360bbf0-bb20-11eb-adfa-773b2eabb1f5',
            'Sec-GPC': '1',
            'Origin': 'https://www.bunnings.co.nz',
            'Sec-Fetch-Site': 'cross-site',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': 'https://www.bunnings.co.nz/',
            'Accept-Language': 'en-US,en;q=0.9',
        }
        response = self.get_json(self.get_response(url, headers, None))
        status = response.get('statusDetails').get('state')
        if status == 'SUCCESS':
            level_1 = response.get('data').get('levels')
            for l1 in level_1[0:3]:
                if running:
                    level_2 = l1.get('levels')
                    for l2 in level_2:      
                        while True and running:
                            tbox_msg = "{} > {}".format(l1.get('displayName'),l2.get('displayName'))
                            TBox.insert(END, '')
                            TBox.insert(END, tbox_msg)
                            url = base_url + l2.get('alternateUrl') + next_page_params.format(page_no=page_no)
                            links = get_products_links(url)
                            page_no += 1
                            if len(links) == 0:
                                break
                            for link in links:
                                if running:
                                    TBox.yview(END)
                                    url = 'https://www.bunnings.co.nz' + link
                                    self.driver.get(url)
                                    try:
                                        myElem = WebDriverWait(self.driver, 120).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.productItemName')))
                                    except TimeoutException:
                                        print("1: Loading took too much time!")
                                    response = self.get_page_source()
                                    data = get_product_details(response, url)
                                    # print(data)
                                    ws.append(data)
                                    tbox_msg = "{}: {}".format(count, data[1])
                                    TBox.insert(END, tbox_msg)
                                    TBox.yview(END)
                                    count += 1
                            wb.save(filename)
                    
        else:
            print('error in getting json..')
        wb.save(filename)
        TBox.insert(END,'Output is saved as {}'.format(filename))
        TBox.yview(END)
        self.driver.quit()

    def mitre10(self):
        global running
        def get_products(product):  
            base_url = 'https://www.mitre10.co.nz'       
            try:
                sku = product.css('.product--sku::text').get().split(':')[1]
            except:
                sku = ''
            productUrl = base_url + product.css('a::attr(href)').get()
            response = self.get_Selector(self.get_response(productUrl,HEADERS,None))
            productCode = response.css('.product--identifiers .product--model-number::text').extract()
            productCode = productCode[0].split(':')[1].strip() if productCode else None
            title = product.css('.product--name::text').get()
            price_list = product.css('.product--dollar-sign::text, .product--price-dollars::text, .product--price-cents::text').extract()
            price = ''.join(price_list)
            # unit = product.css('.product--item-unit::text').get()
            brand = product.css('.product--brand::text').get()
            data = self.clean_data([productCode, title,productUrl, price, brand])
            return data

        filename = self.create_filename('mitre10')
        sheet_title = ['Product Code','Title','Product Url','Price','Brand']
        wb = Workbook()
        ws = wb.active
        ws.append(sheet_title)

        count = 1
        base_url = 'https://www.mitre10.co.nz'
        url = 'https://www.mitre10.co.nz/shop'
        response = self.get_Selector(self.get_response(url, HEADERS, None))
        categoryLinks = response.css('#department-count-component a::attr(href)').extract()
        categories = response.css('#department-count-component a::text').extract()
        i = 0
        for categoryLink in categoryLinks:
            category = categories[i]
            i+=1
            print(category)
            if category == 'Power Tools':
                FOUND = True
            else:
                FOUND = False
            if running and FOUND:
                catUrl = base_url + categoryLink
                response = self.get_Selector(self.get_response(catUrl, HEADERS, None))
                subCategoryLinks = response.css('#department-count-component a::attr(href)').extract()
                subCategories = response.css('#department-count-component a::text').extract()
                j=0
                for subCategoryLink in subCategoryLinks:
                    subCategory = subCategories[j]
                    j+=1
                    if running:
                        tbox_msg = "{} > {}".format(category,subCategory)
                        TBox.insert(END, '')
                        TBox.insert(END, tbox_msg)
                        subCatUrl = base_url + subCategoryLink
                        response = self.get_Selector(self.get_response(subCatUrl, HEADERS, None))
                        productGrid = response.css('.product-grid-item')
                        next_page = response.css('.next a::attr(href)').extract()
                        for product in productGrid:
                            if running:
                                data = get_products(product)
                                print(data)
                                ws.append(data)
                                tbox_msg = "{}: {}".format(count, data[1])
                                TBox.insert(END, tbox_msg)
                                TBox.yview(END)
                                count += 1
                        wb.save(filename)
                        while len(next_page) > 0 and running:
                            nextUrl = subCatUrl + '?' + next_page[0].split('?')[1]
                            response = self.get_Selector(self.get_response(nextUrl, HEADERS, None))
                            productGrid = response.css('.product-grid-item')
                            next_page = response.css('.next a::attr(href)').extract()
                            for product in productGrid:
                                if running:
                                    data = get_products(product)
                                    print(data)
                                    ws.append(data)
                                    tbox_msg = "{}: {}".format(count, data[1])
                                    TBox.insert(END, tbox_msg)
                                    TBox.yview(END)
                                    count += 1
                            wb.save(filename)
            else:
                pass
        wb.save(filename)
        TBox.insert(END,'Output is saved as {}'.format(filename))
        TBox.yview(END)

    def sulco(self):
        global running
        filename = self.create_filename('sulco')
        sheet_title = ['Product Code','Title','Product Url','Special Price','Retail Price']
        wb = Workbook()
        ws = wb.active
        ws.append(sheet_title)
        count = 1

        base_url = 'https://www.sulco.co.nz/'
        catUrl = 'https://www.sulco.co.nz/sulco/site/services/getCategory.ss'
        response = self.get_json(self.get_response(catUrl, HEADERS, None))
        cat_links = [x.get('urlcomponent') for x in response]
        self.open_browser()
        for cat_link in cat_links:
            print(cat_link)
            if running and 'Milwaukee' in cat_link:
                next = True
                self.driver.get(base_url + cat_link)
                try:
                    myElem = WebDriverWait(self.driver, WAIT).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#items .mb-4')))
                    print("Page is ready!")
                except TimeoutException:
                    print("1: Loading took too much time!")
                while next and running:
                    response = self.get_page_source()
                    product_div = response.css('#items .mb-4')
                    for product in product_div:
                        if running:
                            productUrl = base_url[:-1] + product.css('a::attr(href)').get()
                            sku = product.css('.p-1::text').get()
                            title = product.css('.font-size-md::text , .item-code::text').extract()
                            title = ''.join(title)
                            try:
                                retailPrice = product.css('.price-GST::text').get().split(' ')[0]
                            except:
                                retailPrice = ''
                            specialPrice = product.css('.main-prc::text').get()
                            data = self.clean_data([sku, title,productUrl,specialPrice, retailPrice])
                            print(data)
                            ws.append(data)
                            tbox_msg = "{}: {}".format(count, data[1])
                            TBox.insert(END, tbox_msg)
                            TBox.yview(END)
                            count += 1
                    sleep(2)
                    wb.save(filename)
                    try:
                        nextElem = WebDriverWait(self.driver, WAIT).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#next .page-link')))       
                    except TimeoutException:
                        print("2: Loading took too much time!")
                    next_page = self.driver.find_element_by_css_selector('#next .page-link')
                    try:
                        next_page.click()
                        next = True
                    except:
                        next = False
                        print('Khatam tata bye bye')

        self.driver.quit()
        wb.save(filename)
        TBox.insert(END,'Output is saved as {}'.format(filename))
        TBox.yview(END)


    def tradesafety(self):
        global running
        filename = self.create_filename('tradesafety')
        sheet_title = ['Supplier Part No.','Product Code','Title','Product url','Price','Brand']
        wb = Workbook()
        ws = wb.active
        ws.append(sheet_title)
        count = 1

        base_url = 'https://tradesafety.co.nz'

        url = 'https://tradesafety.co.nz/products/browse/catfilter:2.80000092/Safety-Tools/Tools'
        next = True
        self.open_browser()
        # self.driver.get(url)
        # searchBtn = self.driver.find_element_by_css_selector('#catsearch .btn-success')
        # searchBtn.click()
        # sleep(0.5)
        while next and running:
            self.driver.get(url)
            response= self.get_page_source()

            # response = self.get_Selector(self.get_response(url, HEADERS,None))
            product_div = response.css('.col-md-6')
            for product in product_div:
                if running:
                    product_url = product.css('.panel-title a::attr(href)').get()
                    sku = product_url.split('/')[3]
                    title = product.css('.panel-title a::text').get()
                    price = product.css('.clearfix::text').get()
                    brand = product.css('.small::text').get()
                    productUrl = base_url + product_url
                    response1 = self.get_Selector(self.get_response(productUrl,HEADERS,None))
                    table_heads = response1.css('th::text').extract()
                    table_details = response1.css('#maincontainer tr')
                    i = 0
                    for table_head in table_heads:
                        if table_head == 'Supplier part number(s)':
                            sup_part_no = ','.join(table_details[i].css('td::text').extract())
                            break
                        else:
                            pass
                        i+=1
                    data = [sup_part_no,sku, title,productUrl, price, brand]
                    print(data)
                    ws.append(data)
                    tbox_msg = "{}: {}".format(count, data[1])
                    TBox.insert(END, tbox_msg)
                    TBox.yview(END)
                    count += 1
            wb.save(filename)
            sleep(0.5)
            next_pages = response.css('.pagination li')
            for next_page in next_pages:
                if next_page.css('a::attr(title)').get() == 'Next':
                    url = base_url + next_page.css('a::attr(href)').get()
                    next = True
                    break
                else:
                    next = False
        self.driver.quit()
        wb.save(filename)
        TBox.insert(END,'Output is saved as {}'.format(filename))
        TBox.yview(END) 
     
                            
if __name__ == "__main__":
    pp = productPricing()
    root = Tk()
    root.title('Product Pricing Spider')
    icon = PhotoImage(file='./tarantula.png')
    root.iconphoto(False,icon)
    root.geometry('600x400')
    # root.resizable(False, False)
    running = False

    def website_changed():
        global running
        running = True
        website = websitechoosen.get()
        TBox.delete(0,END)
        msg = f'You have selected {website}! Please wait...'
        TBox.insert(END,msg)
        if website == 'https://nzsafetyblackwoods.co.nz':
            pp.nzsafetyblackwoods()
        elif website == 'https://shop.placemakers.co.nz':
            pp.placemakers()
        elif website == 'https://www.hectorjones.co.nz':
            pp.hectorjones()
        elif website == 'https://www.mitre10.co.nz':
            pp.mitre10()
        elif website == 'https://www.bunnings.co.nz':
            pp.bunnings2()
        elif website == 'https://tradesafety.co.nz':
            pp.tradesafety()
        elif website == 'https://www.sulco.co.nz':
            pp.sulco()
        else:
            print('other website is not scripted')
        
    def start():
        # Call work function
        global running
        if running == False:
            t1=Thread(target=website_changed)
            t1.start()
            stopButton.config(text='Stop')
        else:
            TBox.insert(END,'Script is running. Please wait!!!')
    
    def stop():
        global running
        if running:
            if messagebox.askokcancel("Stop", "Do you want to Stop?"):
                running = False
                stopButton.config(text='Quit')
        else:
            if messagebox.askokcancel("Quit", "Do you want to Quit?"):
                root.destroy()

    frame1 = Frame(master=root)
    frame1.pack(fill=X,padx=10, pady=10)
    
    # fontStyle = tkFont.Font(family="Lucida Grande", size=10)
    fontStyle = ("Lucida Grande", 10)
    label1 = Label(frame1, text='Select website: ',font=fontStyle)
    label1.pack(side=LEFT)
    
    n = StringVar()
    websitechoosen = ttk.Combobox(frame1, width = 30,textvariable = n)
    websitechoosen['values'] = ('https://nzsafetyblackwoods.co.nz',
                                'https://shop.placemakers.co.nz',
                                'https://www.hectorjones.co.nz',
                                'https://www.sulco.co.nz',
                                'https://www.bunnings.co.nz',
                                'https://www.mitre10.co.nz', 
                                'https://tradesafety.co.nz'
                                )
    websitechoosen.pack(side=LEFT)
    websitechoosen.current(0)

    stopButton = Button(frame1,text='Quit', width=10, bg='red' ,command=stop)
    stopButton.pack(side=RIGHT)
 
    startButton = Button(frame1,text='Start', width=10, bg='green' ,command=start)
    startButton.pack(side=RIGHT)


  
    frame2 = Frame(root)
    scrollbar = Scrollbar(frame2, orient=VERTICAL, jump=1)
    
    TBox = Listbox(frame2,width=100,height=25,yscrollcommand=scrollbar.set, font=fontStyle)
   
    scrollbar.config(command=TBox.yview)
    scrollbar.pack(side=RIGHT, fill=Y)
    frame2.pack(fill=BOTH, padx=10, pady=10)
    
    TBox.pack()
    
    root.mainloop()


